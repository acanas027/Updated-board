import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter
from io import BytesIO
import os
import shutil
import re
import json
import datetime
from openai import OpenAI


st.set_page_config(page_title="Staffing Report Generator", layout="wide")

st.title("Staffing Report Generator")
st.write("Enter daily inputs, select who is present, and generate the staffing report.")

TEMPLATE_FILE = "staffing_template.xlsx"

if not os.path.exists(TEMPLATE_FILE):
    st.error("Template file not found. Put staffing_template.xlsx in the same folder as report.py.")
    st.stop()


# ============================================================
#  OPPORTUNITY CUSTOMER LIST
# ============================================================
OC_FILE = "Resers DCs Opportunity Cusotmer List.xlsx"
OC_SHEET = "OC Customer List"
OC_HEADER_ROW = 6
OC_DATA_START = 8


@st.cache_data
def load_oc_customer_list():
    if not os.path.exists(OC_FILE):
        st.error(f"OC customer list file not found: '{OC_FILE}'.")
        return []
    try:
        wb = load_workbook(OC_FILE, data_only=True)
        if OC_SHEET not in wb.sheetnames:
            st.error(f"Sheet '{OC_SHEET}' not found in {OC_FILE}.")
            return []
        ws = wb[OC_SHEET]
        customers = []
        for row_idx in range(OC_DATA_START, ws.max_row + 1):
            raw_name = ws.cell(row_idx, 3).value
            if not raw_name:
                continue
            name_clean = str(raw_name).strip().strip('"').lower()
            if "market x" in name_clean or "example" in name_clean:
                continue
            raw_cust_num = ws.cell(row_idx, 2).value
            raw_issue    = ws.cell(row_idx, 5).value
            raw_reqs     = ws.cell(row_idx, 6).value
            raw_signoff  = ws.cell(row_idx, 7).value
            raw_pictures = ws.cell(row_idx, 8).value
            issue = str(raw_issue).strip() if raw_issue else ""
            reqs  = str(raw_reqs).strip()  if raw_reqs  else ""
            sign_off = str(raw_signoff).strip().upper() == "Y" if raw_signoff else False
            pictures = str(raw_pictures).strip().upper() == "Y" if raw_pictures else False
            priority = "HIGH" if (sign_off or pictures) else "MEDIUM"
            base = name_clean.rstrip(" -").split(" - ")[0].strip()
            aliases = []
            for suffix in [" - all loads", " all loads", " fresh dc", " (olathe)"]:
                if base.endswith(suffix):
                    aliases.append(base.replace(suffix, "").strip())
            if "'" in base:
                aliases.append(base.replace("'", ""))
                aliases.append(base.replace("'s", ""))
            known_aliases = {
                "target rialto":          ["target"],
                "sobey's - all loads":    ["sobeys", "sobey", "sobey's"],
                "sysco kc (olathe)":      ["sysco kc", "sysco kansas city", "sysco olathe", "sysco kc olathe"],
                "pfs virgina":            ["pfs virginia", "pfs va"],
                "metro toronto fresh dc": ["metro toronto", "metro fresh"],
                "jewel's":                ["jewels", "jewel"],
                "awg":                    ["associated wholesale grocers"],
                "whataburguer":           ["whataburger"],
            }
            if name_clean in known_aliases:
                aliases += known_aliases[name_clean]
            aliases = list(dict.fromkeys(a for a in aliases if a and a != name_clean))
            customers.append({
                "name": name_clean,
                "aliases": aliases,
                "customer_number": str(raw_cust_num).strip() if raw_cust_num else None,
                "issue": issue,
                "requirements": reqs,
                "sign_off": sign_off,
                "pictures": pictures,
                "priority": priority,
            })
        return customers
    except Exception as e:
        st.error(f"Error loading OC customer list: {e}")
        return []


def find_oc_customers_in_board(board_text):
    oc_list = load_oc_customer_list()
    board_lower = board_text.lower()
    matches = []
    for customer in oc_list:
        search_terms = [customer["name"]] + customer.get("aliases", [])
        found_terms = [term for term in search_terms if term.lower() in board_lower]
        if found_terms:
            matches.append({"customer": customer, "matched_on": found_terms})
    return matches


def build_oc_alert_text(oc_matches):
    if not oc_matches:
        return None
    lines = [
        "=== OPPORTUNITY CUSTOMER (OC) ALERT ===",
        "The following loads belong to customers on the Opportunity Customer List.",
        "These customers have a documented history of complaints and require special handling.",
        "Flag these loads explicitly in your analysis and include action items for each.",
        "",
    ]
    for match in oc_matches:
        c = match["customer"]
        lines.append(f"CUSTOMER: {c['name'].upper()}")
        lines.append(f"Matched on: {', '.join(match['matched_on'])}")
        lines.append(f"Priority: {c['priority']}")
        lines.append(f"Issue History: {c['issue']}")
        lines.append(f"DC Requirements: {c['requirements']}")
        if c["sign_off"]:
            lines.append("DC Supervisor Sign-Off REQUIRED before this load ships.")
        if c["pictures"]:
            lines.append("Photos REQUIRED: 3 on dock + 3 during loading (6 total). Email to manager.")
        lines.append("")
    lines += [
        "IMPORTANT: For every OC load identified above:",
        "1. Flag it clearly in your Board Summary section.",
        "2. Add a dedicated OC Action section with specific steps before this load ships.",
        "3. Recommend who should own the sign-off and photo process.",
        "4. Include this as one of the Top 3 Action Items if the load is active or upcoming.",
        "=== END OC ALERT ===",
    ]
    return "\n".join(lines)


def get_groq_client():
    if "GROQ_API_KEY" not in st.secrets:
        return None
    return OpenAI(
        api_key=st.secrets["GROQ_API_KEY"],
        base_url="https://api.groq.com/openai/v1",
    )


@st.cache_data
def load_names_for_shift(shift):
    wb = load_workbook(TEMPLATE_FILE, data_only=True)
    if shift == "1st":
        ws = wb["Staffing sheet 1ST Shift"]
    else:
        ws = wb["Staffing Sheet 2nd Shift"]
    names = []
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row, 1).value
        if val and str(val).strip():
            names.append(str(val).strip())
        elif names:
            break
    return sorted(names)


def whole_workers(value):
    return int(float(value or 0) + 0.7)


def is_present(row):
    return str(row["Present"]).strip().lower() == "x"


def has_skill(row, code):
    return code in str(row["Skills"])


def best_fit(row, text):
    return text.lower() in str(row["Best Fit"]).lower()


def name_contains(row, text):
    return text.lower() in str(row["Name"]).lower()


def calculate_input_values(day, shift, total_cases):
    first_shift_pick = {
        "Sunday": 0.20, "Monday": 0.18, "Tuesday": 0.18, "Wednesday": 0.19,
        "Thursday": 0.19, "Friday": 0.18, "Saturday": 0.21,
    }
    second_shift_pick = {
        "Sunday": 0.19, "Monday": 0.15, "Tuesday": 0.15, "Wednesday": 0.17,
        "Thursday": 0.17, "Friday": 0.17, "Saturday": 0.19,
    }
    first_shift_fp = {
        "Sunday": 0.28, "Monday": 0.32, "Tuesday": 0.40, "Wednesday": 0.35,
        "Thursday": 0.35, "Friday": 0.36, "Saturday": 0.31,
    }
    second_shift_fp = {
        "Sunday": 0.32, "Monday": 0.33, "Tuesday": 0.27, "Wednesday": 0.29,
        "Thursday": 0.28, "Friday": 0.30, "Saturday": 0.30,
    }
    if shift == "1st":
        cases_to_pick = total_cases * first_shift_pick.get(day, 0)
        full_pallets = (total_cases * first_shift_fp.get(day, 0)) / 70
    else:
        cases_to_pick = total_cases * second_shift_pick.get(day, 0)
        full_pallets = (total_cases * second_shift_fp.get(day, 0)) / 70
    return cases_to_pick, full_pallets


def calculate_needed(
    day, shift, total_cases, hours_remaining, total_outbound_loads_actual,
    crossroads_open, deer_creek_open, msb_open,
):
    if hours_remaining <= 0:
        hours_remaining = 1
    cases_to_pick, full_pallets = calculate_input_values(day, shift, total_cases)
    inbound_pallets = 0
    if crossroads_open == "YES":
        inbound_pallets += 700
    if deer_creek_open == "YES":
        inbound_pallets += 500
    if msb_open == "YES":
        inbound_pallets += 640
    raw_needed = {
        "Unloading":     (inbound_pallets / 4) / (44 * hours_remaining),
        "Receiving":     (inbound_pallets / 4) / (44 * hours_remaining),
        "Putaway":       (inbound_pallets / 2) / (25 * hours_remaining),
        "Picking":       cases_to_pick / (185 * hours_remaining),
        "Replenishment": (cases_to_pick / 70) / (25 * 8.5),
        "Full Pallets":  full_pallets / (25 * hours_remaining),
        "Loading":       total_outbound_loads_actual / hours_remaining,
    }
    needed = {
        "Unloading": max(2, whole_workers(raw_needed["Unloading"])),
        "Receiving":  max(2, whole_workers(raw_needed["Receiving"])),
        "Picking":    whole_workers(raw_needed["Picking"]),
        "Tasking":    whole_workers(
            raw_needed["Putaway"] + raw_needed["Replenishment"] + raw_needed["Full Pallets"]
        ),
        "Loading":    whole_workers(raw_needed["Loading"]),
    }
    return needed, raw_needed, cases_to_pick, full_pallets, inbound_pallets


def generate_recommendations(staff, needed):
    assigned = {task: 0 for task in needed}
    staff["Recommended Task"] = ""
    present_indexes = staff[staff.apply(is_present, axis=1)].index.tolist()

    def assign_if_needed(task, idx):
        if assigned[task] < needed[task]:
            staff.at[idx, "Recommended Task"] = task
            assigned[task] += 1
            return True
        return False

    for idx in present_indexes:
        row = staff.loc[idx]
        if name_contains(row, "Dale"):
            staff.at[idx, "Recommended Task"] = "Receiving"
            assigned["Receiving"] += 1
        elif name_contains(row, "Alex"):
            staff.at[idx, "Recommended Task"] = "Unloading"
            assigned["Unloading"] += 1

    for idx in present_indexes:
        if staff.at[idx, "Recommended Task"] != "":
            continue
        row = staff.loc[idx]
        if str(row["Skills"]).strip() == "P":
            assign_if_needed("Picking", idx)

    best_fit_steps = [
        ("Unloading", "Unload", "U"),
        ("Loading",   "Load",   "L"),
        ("Receiving", "Receiv", "R"),
        ("Picking",   "Pick",   "P"),
        ("Tasking",   "Task",   "T"),
    ]
    for task, fit_text, skill in best_fit_steps:
        for idx in present_indexes:
            if staff.at[idx, "Recommended Task"] != "":
                continue
            row = staff.loc[idx]
            if best_fit(row, fit_text) and has_skill(row, skill):
                assign_if_needed(task, idx)

    skill_map = {"Unloading": "U", "Receiving": "R", "Loading": "L", "Picking": "P", "Tasking": "T"}
    for task, skill in skill_map.items():
        for idx in present_indexes:
            if assigned[task] >= needed[task]:
                break
            if staff.at[idx, "Recommended Task"] != "":
                continue
            row = staff.loc[idx]
            if has_skill(row, skill):
                assign_if_needed(task, idx)

    backup_tasks = ["Unloading", "Receiving", "Loading", "Picking", "Tasking"]
    for task in backup_tasks:
        while assigned[task] < needed[task]:
            found_worker = False
            for idx in present_indexes:
                if staff.at[idx, "Recommended Task"] != "":
                    continue
                row = staff.loc[idx]
                if best_fit(row, "Task") and (has_skill(row, "T") or has_skill(row, "L") or has_skill(row, "P")):
                    assign_if_needed(task, idx)
                    found_worker = True
                    break
            if found_worker:
                continue
            for idx in present_indexes:
                if staff.at[idx, "Recommended Task"] != "":
                    continue
                row = staff.loc[idx]
                if has_skill(row, "T") or has_skill(row, "L") or has_skill(row, "P"):
                    assign_if_needed(task, idx)
                    found_worker = True
                    break
            if not found_worker:
                break

    for idx in present_indexes:
        if staff.at[idx, "Recommended Task"] == "":
            if assigned["Tasking"] < needed["Tasking"]:
                staff.at[idx, "Recommended Task"] = "Tasking"
                assigned["Tasking"] += 1
            else:
                staff.at[idx, "Recommended Task"] = "Lead/Extra"

    preferred_extra_names = ["will", "antonio"]
    preferred_idxs = [
        idx for idx in present_indexes
        if any(name in str(staff.at[idx, "Name"]).lower() for name in preferred_extra_names)
    ]
    current_extra_idxs = [
        idx for idx in present_indexes
        if staff.at[idx, "Recommended Task"] == "Lead/Extra"
    ]
    for preferred_idx in preferred_idxs:
        if not current_extra_idxs:
            break
        if staff.at[preferred_idx, "Recommended Task"] == "Lead/Extra":
            continue
        swap_idx = None
        for extra_idx in current_extra_idxs:
            if not any(name in str(staff.at[extra_idx, "Name"]).lower() for name in preferred_extra_names):
                swap_idx = extra_idx
                break
        if swap_idx is None:
            break
        old_task = staff.at[preferred_idx, "Recommended Task"]
        staff.at[preferred_idx, "Recommended Task"] = "Lead/Extra"
        staff.at[swap_idx, "Recommended Task"] = old_task
        current_extra_idxs.remove(swap_idx)

    return staff


def build_summary(staff, needed):
    present_recommendations = staff[
        staff["Present"].astype(str).str.strip().str.lower().eq("x")
        & staff["Recommended Task"].astype(str).str.strip().ne("")
    ].copy()
    needed_list   = pd.Series(needed, name="Needed")
    assigned_list = present_recommendations["Recommended Task"].value_counts().rename("Assigned")
    summary_table = pd.concat([needed_list, assigned_list], axis=1).fillna(0)
    summary_table["Needed"]     = summary_table["Needed"].astype(int)
    summary_table["Assigned"]   = summary_table["Assigned"].astype(int)
    summary_table["Difference"] = summary_table["Assigned"] - summary_table["Needed"]
    summary_table["Status"]     = summary_table["Difference"].apply(
        lambda x: "Good" if x == 0 else ("Overstaffed" if x > 0 else "Understaffed")
    )
    return present_recommendations, summary_table


def build_recommendations(summary_table, present_recommendations, raw_needed, hours_remaining, notes):
    recommendations = []
    total_labor_gap = int(summary_table["Difference"].sum())
    labor_hours_gap = total_labor_gap * hours_remaining
    recommendations.append(
        f"Current labor balance estimate: {labor_hours_gap:+.1f} labor-hours. "
        f"Positive means extra capacity; negative means short capacity."
    )
    for task, row in summary_table.iterrows():
        diff = int(row["Difference"])
        if diff < 0:
            recommendations.append(
                f"{task}: approximately {abs(diff * hours_remaining):.1f} labor-hours behind based on current staffing vs need."
            )
        elif diff > 0:
            recommendations.append(f"{task}: approximately {diff * hours_remaining:.1f} labor-hours ahead / available capacity.")
        else:
            recommendations.append(f"{task}: Staffing is balanced.")

    picking_gap   = int(summary_table.loc["Picking",    "Difference"]) if "Picking"    in summary_table.index else 0
    tasking_gap   = int(summary_table.loc["Tasking",    "Difference"]) if "Tasking"    in summary_table.index else 0
    receiving_gap = int(summary_table.loc["Receiving",  "Difference"]) if "Receiving"  in summary_table.index else 0
    unloading_gap = int(summary_table.loc["Unloading",  "Difference"]) if "Unloading"  in summary_table.index else 0
    loading_gap   = int(summary_table.loc["Loading",    "Difference"]) if "Loading"    in summary_table.index else 0
    lead_gap      = int(summary_table.loc["Lead/Extra", "Difference"]) if "Lead/Extra" in summary_table.index else 0

    if picking_gap < 0:
        recommendations.append("High picking short risk detected. Consider moving tasking labor into replenishment to protect pickers.")
        recommendations.append("Avoid pulling pickers into unloading or loading unless outbound service is critical.")
        if tasking_gap > 0:
            recommendations.append(f"Tasking currently has {tasking_gap} extra worker(s). Consider temporarily assigning them to replenishment.")
        if lead_gap > 0:
            recommendations.append("Lead/Extra capacity available. Consider flexing extra labor into replenishment or picking support.")
    if unloading_gap < 0 or receiving_gap < 0:
        recommendations.append("Inbound flow risk detected. Falling behind may create dock congestion and delayed putaway.")
        recommendations.append("Consider moving flexible tasking labor into unloading or receiving temporarily.")
        if tasking_gap > 1:
            recommendations.append("Tasking has available labor that can support inbound operations.")
    if loading_gap < 0:
        recommendations.append("Outbound loading risk detected. Late departures and service failures may increase.")
        recommendations.append("Protect loading labor before reallocating to non-critical work.")
        if lead_gap > 0:
            recommendations.append("Use Lead/Extra labor to support outbound staging or trailer cleanup.")
    if total_labor_gap > 1:
        recommendations.append("Operation currently has excess labor capacity.")
        recommendations.append("Consider deep cleaning, trailer audits, replenishment cleanup, or cross-training.")
        recommendations.append("Extra labor could be used proactively to prevent later picking shortages.")

    inbound_pressure  = raw_needed["Unloading"] + raw_needed["Receiving"] + raw_needed["Putaway"]
    outbound_pressure = raw_needed["Picking"] + raw_needed["Loading"]
    if inbound_pressure > outbound_pressure * 1.3:
        recommendations.append("Inbound workload is significantly heavier than outbound.")
        recommendations.append("Focus on unloading, receiving, and putaway to avoid congestion.")
    elif outbound_pressure > inbound_pressure * 1.3:
        recommendations.append("Outbound workload is significantly heavier than inbound.")
        recommendations.append("Prioritize replenishment and picking continuity to avoid shorts.")

    if hours_remaining <= 4:
        recommendations.append("Shift is entering final hours. Prioritize completion work and outbound execution.")
    elif hours_remaining >= 8:
        recommendations.append("Enough shift time remains to strategically rebalance labor before bottlenecks form.")

    lower_notes = notes.lower()
    if "late"  in lower_notes:
        recommendations.append("Manager notes mention late loads. Prioritize outbound execution and trailer readiness.")
    if "short" in lower_notes:
        recommendations.append("Manager notes indicate short risk. Protect replenishment and picking flow.")
    if "live"  in lower_notes:
        recommendations.append("Live loads detected in notes. Prioritize those doors before drop trailers.")
    if "cpu"   in lower_notes:
        recommendations.append("CPU loads referenced. Ensure loading labor is protected.")

    return recommendations


# ============================================================
#  BOARD EXCEL READING
# ============================================================
BOARD_DAY_NAMES = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]


def normalize_board_text(value):
    if value is None:
        return ""
    if isinstance(value, datetime.time):
        return value.strftime("%H:%M")
    if isinstance(value, datetime.datetime):
        return value.strftime("%m/%d/%Y")
    try:
        import math
        if isinstance(value, float) and math.isnan(value):
            return ""
    except Exception:
        pass
    text = str(value).replace("\n", " ").strip()
    if text.endswith(".0"):
        text = text[:-2]
    return text


def normalize_board_date(value):
    text = normalize_board_text(value)
    if not text:
        return ""
    for fmt in ["%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"]:
        try:
            return pd.to_datetime(text, format=fmt).strftime("%m/%d/%Y")
        except Exception:
            pass
    try:
        parsed = pd.to_datetime(text, errors="coerce")
        if pd.notna(parsed):
            return parsed.strftime("%m/%d/%Y")
    except Exception:
        pass
    return text


def normalize_board_time(value):
    text = normalize_board_text(value)
    if not text:
        return ""
    if re.fullmatch(r"\d{1,2}:\d{2}", text):
        h, m = text.split(":")
        return f"{int(h):02d}:{m}"
    try:
        parsed = pd.to_datetime(text, errors="coerce")
        if pd.notna(parsed):
            return parsed.strftime("%H:%M")
    except Exception:
        pass
    return text


def looks_like_board_load(value):
    text = normalize_board_text(value)
    if not text:
        return ""
    if text in BOARD_DAY_NAMES:
        return ""
    if re.search(r"\d{1,2}/\d{1,2}/\d{2,4}", text):
        return ""
    if re.fullmatch(r"\d{1,2}:\d{2}", text):
        return ""
    digits = re.sub(r"[^0-9]", "", text)
    if 5 <= len(digits) <= 9:
        return digits
    return ""


def detect_board_status(value):
    text = normalize_board_text(value).upper()
    if not text:
        return ""
    if "LOADED SHORT" in text:
        return "Loaded Short"
    if "PICKING/SHORT" in text or "PICKING SHORT" in text:
        return "Picking/Short"
    if "READY/SHORT" in text or re.search(r"\bR/S\b", text):
        return "R/S"
    if re.search(r"\bRTL\b", text) or "READY TO LOAD" in text:
        return "RTL"
    if "NO DRIVER" in text:
        return "No Driver"
    if "PICKING" in text:
        return "Picking"
    if "COMPLETED" in text or re.search(r"\bCOMPLETE\b", text):
        return "Completed"
    if re.search(r"\bLATE\b", text):
        return "Late"
    if re.search(r"\bLOADED\b", text):
        return "Loaded"
    return ""


def detect_trailer_field_late(trailer_value):
    text = normalize_board_text(trailer_value).upper()
    if not text:
        return False
    if re.search(r"\bLATE\b", text):
        return True
    if re.match(r"^ETA\b", text):
        return True
    return False


def board_cell_flags(cell):
    flags = []
    fill_color = ""
    font_color = ""
    try:
        fill = cell.fill
        if fill and fill.fgColor:
            if fill.fgColor.type == "rgb":
                fill_color = str(fill.fgColor.rgb).upper()
            elif fill.fgColor.type == "indexed":
                fill_color = str(fill.fgColor.indexed).upper()
    except Exception:
        pass
    try:
        font = cell.font
        if font and font.color and font.color.type == "rgb":
            font_color = str(font.color.rgb).upper()
    except Exception:
        pass
    if fill_color in ("FFFFFF00", "00FFFF00", "FFFF00", "0000000D"):
        flags.append("LOAD-CHECK")
    if fill_color in ("FFADD8E6", "FF87CEEB", "FFADD8FF", "FFB0E0E6", "FF00BFFF"):
        flags.append("TT4-NEEDED")
    if font_color in ("FFFF0000", "00FF0000"):
        flags.append("CANADIAN")
    return flags


def parse_number(value):
    text = normalize_board_text(value)
    if not text or text.strip() in ("", " "):
        return 0
    digits = re.sub(r"[^0-9]", "", text)
    return int(digits) if digits else 0


# ============================================================
#  READ BOARD-LEVEL TOTALS FROM ROW 2
#  K2 = total pulls remaining on the board right now
#  L2 = total picks remaining on the board right now
#  These are live SUM formulas. When a load goes RTL or
#  Completed, the crew erases its K/L values, so these totals
#  always reflect actual remaining work — not projected work.
# ============================================================
def read_board_totals(board_file):
    """
    Read K2 (total pulls remaining) and L2 (total picks remaining)
    directly from the board. Returns (pulls_remaining, picks_remaining).
    Falls back to (None, None) if the file can't be read or cells are empty.
    """
    try:
        board_file.seek(0)
        if board_file.name.lower().endswith(".csv"):
            return None, None
        wb = load_workbook(board_file, data_only=True)
        outbound_sheet = None
        for candidate in ["Outbound", "outbound", "OUTBOUND"]:
            if candidate in wb.sheetnames:
                outbound_sheet = candidate
                break
        if not outbound_sheet:
            return None, None
        ws = wb[outbound_sheet]
        pulls = ws.cell(2, 11).value  # K2
        picks = ws.cell(2, 12).value  # L2
        pulls = int(pulls) if pulls and str(pulls).strip() not in ("", "0") else None
        picks = int(picks) if picks and str(picks).strip() not in ("", "0") else None
        return pulls, picks
    except Exception:
        return None, None


def board_records_from_excel(board_file):
    board_file.seek(0)
    file_name = board_file.name.lower()
    all_rows = []

    if file_name.endswith(".xls"):
        sheets = pd.read_excel(board_file, sheet_name=None, header=None, engine="xlrd")
        for sheet_name, df in sheets.items():
            df = df.fillna("")
            current_day = ""
            current_date = ""
            for idx, row in df.iterrows():
                values = [normalize_board_text(v) for v in row.tolist()]
                while len(values) < 13:
                    values.append("")
                first_cell = values[0]
                if first_cell in BOARD_DAY_NAMES:
                    current_day = first_cell
                    current_date = normalize_board_date(values[1])
                    continue
                load_number = looks_like_board_load(values[0])
                if not load_number:
                    continue
                if detect_trailer_field_late(values[6]):
                    status = "Late"
                else:
                    status = detect_board_status(values[7])
                if not status:
                    status = detect_board_status(" ".join(values))
                # Col D (values[3]) = Type — human-typed, reliable (CPU, CPU-Drop, CPU-Live, Live, Drop)
                # Col G (values[6]) = Trailer — XLOOKUP to external file, often blank; used as fallback
                type_raw = values[3].upper()
                trailer_text = values[6].upper()
                if type_raw:
                    type_value = values[3]  # preserve original casing from col D
                else:
                    type_value = "Live" if "LIVE" in trailer_text else ("CPU - Live" if "CPU" in trailer_text else ("Drop" if "DROP" in trailer_text else ""))
                all_rows.append({
                    "source":      sheet_name,
                    "day":         current_day,
                    "date":        current_date,
                    "load_number": load_number,
                    "customer":    values[1],
                    "carrier":     values[2],
                    "appt_time":   normalize_board_time(values[4]),
                    "door":        values[5],
                    "trailer":     values[6],
                    "status":      status,
                    "type":        type_value,
                    "tt4":         values[8],
                    "loader":      values[9],
                    "comments":    values[14] if len(values) > 14 else "",
                    "pulls":       parse_number(values[10]),
                    "picks":       parse_number(values[11]),
                    "flags":       [],
                    "raw_row":     " | ".join(v for v in values if v),
                })
        return all_rows

    wb = load_workbook(board_file, data_only=True)
    outbound_sheet = None
    for candidate in ["Outbound", "outbound", "OUTBOUND"]:
        if candidate in wb.sheetnames:
            outbound_sheet = candidate
            break
    sheets_to_read = [outbound_sheet] if outbound_sheet else wb.sheetnames
    for sheet_name in sheets_to_read:
        ws = wb[sheet_name]
        current_day = ""
        current_date = ""
        consecutive_empty = 0
        for row_idx in range(1, ws.max_row + 1):
            values = []
            flags = []
            has_content = False
            for col_idx in range(1, 14):
                cell = ws.cell(row_idx, col_idx)
                if cell.value is not None:
                    has_content = True
                values.append(normalize_board_text(cell.value))
                for flag in board_cell_flags(cell):
                    flags.append(flag)

            if not has_content:
                consecutive_empty += 1
                if consecutive_empty >= 15:
                    break
                continue
            consecutive_empty = 0

            first_cell = values[0]
            if first_cell in BOARD_DAY_NAMES:
                current_day = first_cell
                current_date = normalize_board_date(values[1])
                continue

            load_number = looks_like_board_load(values[0])
            if not load_number:
                continue

            if detect_trailer_field_late(values[6]):
                status = "Late"
            else:
                status = detect_board_status(values[7])
            if not status:
                status = detect_board_status(" ".join(values))

            # Col D (values[3]) = Type — human-typed, reliable
            # Col G (values[6]) = Trailer — fallback only
            type_raw = values[3].upper()
            trailer_text = values[6].upper()
            if type_raw:
                type_value = values[3]  # preserve original casing from col D
            else:
                type_value = "Live" if "LIVE" in trailer_text else ("CPU - Live" if "CPU" in trailer_text else ("Drop" if "DROP" in trailer_text else ""))

            all_rows.append({
                "source":      sheet_name,
                "row_number":  row_idx,
                "day":         current_day,
                "date":        current_date,
                "load_number": load_number,
                "customer":    values[1],
                "carrier":     values[2],
                "appt_time":   normalize_board_time(values[4]),
                "door":        values[5],
                "trailer":     values[6],
                "status":      status,
                "type":        type_value,
                "tt4":         values[8],
                "loader":      values[9],
                "comments":    values[14] if len(values) > 14 else "",
                "pulls":       parse_number(values[10]),
                "picks":       parse_number(values[11]),
                "flags":       sorted(set(flags)),
                "raw_row":     " | ".join(v for v in values if v),
            })

    return all_rows


def board_records_from_csv(board_file):
    board_file.seek(0)
    df = pd.read_csv(board_file, header=None).fillna("")
    current_day = ""
    current_date = ""
    all_rows = []
    for idx, row in df.iterrows():
        values = [normalize_board_text(v) for v in row.tolist()]
        while len(values) < 13:
            values.append("")
        first_cell = values[0]
        if first_cell in BOARD_DAY_NAMES:
            current_day = first_cell
            current_date = normalize_board_date(values[1])
            continue
        load_number = looks_like_board_load(values[0])
        if not load_number:
            continue
        if detect_trailer_field_late(values[6]):
            status = "Late"
        else:
            status = detect_board_status(values[7])
        if not status:
            status = detect_board_status(" ".join(values))
        # Col D (values[3]) = Type — human-typed, reliable
        type_raw = values[3].upper()
        trailer_text = values[6].upper()
        if type_raw:
            type_value = values[3]
        else:
            type_value = "Live" if "LIVE" in trailer_text else ("CPU - Live" if "CPU" in trailer_text else ("Drop" if "DROP" in trailer_text else ""))
        all_rows.append({
            "source":      "CSV Board",
            "day":         current_day,
            "date":        current_date,
            "load_number": load_number,
            "customer":    values[1],
            "carrier":     values[2],
            "appt_time":   normalize_board_time(values[4]),
            "door":        values[5],
            "trailer":     values[6],
            "status":      status,
            "type":        type_value,
            "tt4":         values[8],
            "loader":      values[9],
            "comments":    values[14] if len(values) > 14 else "",
            "pulls":       parse_number(values[10]),
            "picks":       parse_number(values[11]),
            "flags":       [],
            "raw_row":     " | ".join(v for v in values if v),
        })
    return all_rows


def board_records_from_inbound_sheet(board_file):
    board_file.seek(0)
    try:
        wb = load_workbook(board_file, data_only=True)
    except Exception:
        return []

    inbound_sheet = None
    for candidate in ["Inbound", "inbound", "INBOUND"]:
        if candidate in wb.sheetnames:
            inbound_sheet = candidate
            break
    if not inbound_sheet:
        return []

    ws = wb[inbound_sheet]
    all_rows = []
    current_day = ""
    current_date = ""

    for row_idx in range(1, ws.max_row + 1):
        has_content = any(ws.cell(row_idx, c).value is not None for c in range(1, 12))
        if not has_content:
            continue
        col1 = normalize_board_text(ws.cell(row_idx, 1).value)
        col1_title = col1.strip().title()
        if col1_title in BOARD_DAY_NAMES:
            current_day = col1_title
            current_date = normalize_board_date(ws.cell(row_idx, 2).value)
            continue
        if col1.lower() in ("load number", "load #", "load"):
            continue
        load_number = looks_like_board_load(col1)
        if not load_number:
            continue
        all_rows.append({
            "source":      inbound_sheet,
            "day":         current_day,
            "date":        current_date,
            "load_number": load_number,
            "carrier":     normalize_board_text(ws.cell(row_idx, 2).value),
            "appt_time":   normalize_board_time(ws.cell(row_idx, 3).value),
            "type":        normalize_board_text(ws.cell(row_idx, 4).value),
            "trailer":     normalize_board_text(ws.cell(row_idx, 5).value),
            "status":      normalize_board_text(ws.cell(row_idx, 6).value),
            "receiver":    normalize_board_text(ws.cell(row_idx, 7).value),
            "origin":      normalize_board_text(ws.cell(row_idx, 8).value),
            "or_number":   normalize_board_text(ws.cell(row_idx, 9).value),
            "notes":       normalize_board_text(ws.cell(row_idx, 10).value),
        })

    return all_rows


def build_python_inbound_summary(inbound_rows):
    summary = {
        "loads_read_from_inbound": len(inbound_rows),
        "loads_by_day": {},
        "live_loads": 0,
        "drop_loads": 0,
        "on_lot": 0,
        "at_door": 0,
        "loads_with_receiver": 0,
        "loads_missing_receiver": 0,
    }
    for row in inbound_rows:
        day_key = row.get("day") or "Unknown Day"
        summary["loads_by_day"][day_key] = summary["loads_by_day"].get(day_key, 0) + 1
        type_upper   = row.get("type",   "").upper()
        status_upper = row.get("status", "").upper()
        if "LIVE"   in type_upper: summary["live_loads"] += 1
        if "DROP"   in type_upper: summary["drop_loads"] += 1
        if "ON LOT" in status_upper: summary["on_lot"] += 1
        if "DOOR"   in status_upper: summary["at_door"] += 1
        if row.get("receiver"): summary["loads_with_receiver"] += 1
        else: summary["loads_missing_receiver"] += 1
    return summary


def build_python_board_summary(board_rows):
    summary = {
        "loads_read_from_board":        len(board_rows),
        "loads_by_day":                 {},
        "loads_by_date":                {},
        "status_counts":                {},
        "late_loads":                   0,
        "rtl_loads":                    0,
        "rs_loads":                     0,
        "picking_loads":                0,
        "picking_short_loads":          0,
        "loaded_short_loads":           0,
        "completed_loads":              0,
        "blank_or_not_started_loads":   0,
        "live_loads":                   0,
        "drop_loads":                   0,
        "cpu_loads":                    0,
        "tt4_needed_loads":             0,
        "load_check_loads":             0,
        "canadian_loads":               0,
        "loads_with_loader_assigned":   0,
        "loads_missing_loader":         0,
        "late_load_details":            [],
        "rs_load_details":              [],
        "picking_short_details":        [],
        "loaded_short_details":         [],
        "rtl_details":                  [],
        "blank_or_not_started_details": [],
        "priority_load_details":        [],
    }

    for row in board_rows:
        day_key      = row.get("day")    or "Unknown Day"
        date_key     = row.get("date")   or "Unknown Date"
        status       = row.get("status") or "Blank/Not Started"
        status_upper = status.upper()
        raw_upper    = row.get("raw_row", "").upper()
        flags        = row.get("flags", [])

        summary["loads_by_day"][day_key]   = summary["loads_by_day"].get(day_key, 0)   + 1
        summary["loads_by_date"][date_key] = summary["loads_by_date"].get(date_key, 0) + 1
        summary["status_counts"][status]   = summary["status_counts"].get(status, 0)   + 1

        if "LATE" in status_upper or "LATE " in f" {raw_upper} ":
            summary["late_loads"] += 1
            summary["late_load_details"].append(row)
        if status_upper == "RTL" or "READY TO LOAD" in status_upper:
            summary["rtl_loads"] += 1
            summary["rtl_details"].append(row)
        if status_upper in ["R/S", "READY/SHORT"] or "R/S" in raw_upper:
            summary["rs_loads"] += 1
            summary["rs_load_details"].append(row)
        if status_upper == "PICKING":
            summary["picking_loads"] += 1
        if "PICKING/SHORT" in status_upper or "PICKING SHORT" in status_upper:
            summary["picking_short_loads"] += 1
            summary["picking_short_details"].append(row)
        if "LOADED SHORT" in status_upper:
            summary["loaded_short_loads"] += 1
            summary["loaded_short_details"].append(row)
        if "COMPLETED" in status_upper or status_upper == "COMPLETE":
            summary["completed_loads"] += 1
        if not row.get("status"):
            summary["blank_or_not_started_loads"] += 1
            summary["blank_or_not_started_details"].append(row)
        if "LIVE" in raw_upper:
            summary["live_loads"] += 1
            summary["priority_load_details"].append(row)
        if "DROP" in raw_upper:
            summary["drop_loads"] += 1
        if "CPU" in raw_upper:
            summary["cpu_loads"] += 1
            summary["priority_load_details"].append(row)
        if "TT4-NEEDED" in flags:
            summary["tt4_needed_loads"] += 1
            summary["priority_load_details"].append(row)
        if "LOAD-CHECK" in flags:
            summary["load_check_loads"] += 1
            summary["priority_load_details"].append(row)
        if "CANADIAN" in flags:
            summary["canadian_loads"] += 1
            summary["priority_load_details"].append(row)
        if row.get("loader"):
            summary["loads_with_loader_assigned"] += 1
        else:
            summary["loads_missing_loader"] += 1

    seen = set()
    unique_priority = []
    for item in summary["priority_load_details"]:
        key = (item.get("load_number"), item.get("row_number"), item.get("source"))
        if key not in seen:
            seen.add(key)
            unique_priority.append(item)
    summary["priority_load_details"] = unique_priority

    return summary


def slim_summary_for_ai(board_summary):
    detail_keys = {k for k in board_summary if k.endswith("_details")}
    return {k: v for k, v in board_summary.items() if k not in detail_keys}


def actionable_rows_for_ai(board_rows):
    """
    Two buckets:
    - actionable: loads needing attention. Includes pulls and picks so the AI
      knows exactly how much work each load still requires.
    - completed: slim records for pacing context only.
    RTL loads are included in actionable so the AI can see what is staged
    and ready, and reference their remaining pulls/picks (should be 0 if
    the crew has already erased them per workflow).
    """
    COMPLETED_STATUSES = {"Completed", "Complete"}
    SKIP_STATUSES = {"Loaded"}
    actionable = []
    completed = []
    for row in board_rows:
        status = (row.get("status") or "").strip()
        flags  = row.get("flags", [])
        is_blank       = not status
        is_completed   = status in COMPLETED_STATUSES
        is_skip        = status in SKIP_STATUSES and not flags
        notable_status = status and not is_completed and not is_skip

        if is_completed:
            completed.append({
                "day":      row.get("day", ""),
                "load":     row.get("load_number", ""),
                "customer": row.get("customer", ""),
                "time":     row.get("appt_time", ""),
                "status":   status,
            })
        elif notable_status or bool(flags) or is_blank:
            actionable.append({
                "day":      row.get("day", ""),
                "load":     row.get("load_number", ""),
                "customer": row.get("customer", ""),
                "time":     row.get("appt_time", ""),
                "door":     row.get("door", ""),
                "trailer":  row.get("trailer", ""),
                "status":   status or "Blank/Not Started",
                "type":     row.get("type", ""),
                "loader":   row.get("loader", ""),
                "pulls":    row.get("pulls", 0),   # actual remaining pulls on this load
                "picks":    row.get("picks", 0),   # actual remaining picks on this load
                "flags":    flags,
                "comments": row.get("comments", ""),
            })
    return actionable, completed


def read_board_file_to_text(board_file):
    board_file.seek(0)
    file_name = board_file.name.lower()

    try:
        if file_name.endswith(".csv"):
            board_rows   = board_records_from_csv(board_file)
            inbound_rows = []
        else:
            board_rows = board_records_from_excel(board_file)
            board_file.seek(0)
            inbound_rows = board_records_from_inbound_sheet(board_file)

        board_summary   = build_python_board_summary(board_rows)
        inbound_summary = build_python_inbound_summary(inbound_rows)
        actionable_rows, completed_rows = actionable_rows_for_ai(board_rows)

        payload = {
            "python_verified_outbound_summary": slim_summary_for_ai(board_summary),
            "python_verified_inbound_summary":  inbound_summary,
            "actionable_outbound_rows":         actionable_rows,
            "completed_outbound_rows":          completed_rows,
            "instructions_for_ai": [
                "Use python_verified_outbound_summary for ALL outbound counts — do not recount from rows.",
                "Use python_verified_inbound_summary for ALL inbound counts.",
                "actionable_outbound_rows = loads needing attention (notable status, flagged, or blank). Each row includes pulls and picks showing actual remaining work for that load.",
                "completed_outbound_rows = slim records of finished loads. Use appt times to judge pacing.",
                "pulls and picks per load are ACTUAL remaining values. When a load goes RTL or Completed the crew erases its pulls/picks, so 0 means done. Non-zero means work still needed.",
                "Outbound and inbound are separate — never mix their counts.",
                "All times use 24-hour clock.",
                "Blank status means load not yet started.",
                "Flags: LOAD-CHECK=yellow fill, TT4-NEEDED=blue fill, CANADIAN=red font.",
                "RTL and Completed loads are already picked/pulled. Do NOT count them as remaining work.",
            ],
        }

        return json.dumps(payload, indent=2, ensure_ascii=False)

    except Exception as e:
        error_message = str(e)
        st.error(f"BOARD PARSER ERROR: {error_message}")
        st.exception(e)
        return json.dumps({
            "error": f"Could not read board file: {error_message}",
            "python_verified_outbound_summary": {},
            "python_verified_inbound_summary":  {},
            "actionable_outbound_rows": [],
        }, indent=2, ensure_ascii=False)


# ============================================================
#  SINGLE-CALL GROQ ANALYSIS
# ============================================================
def _rows_to_table(rows, columns):
    if not rows:
        return "(none)"
    header = " | ".join(columns)
    sep    = "-" * len(header)
    lines  = [header, sep]
    for r in rows:
        lines.append(" | ".join(str(r.get(c, "")).strip() for c in columns))
    return "\n".join(lines)


def analyze_board_with_groq(
    board_text, day, shift, total_cases, hours_remaining, total_outbound_loads,
    crossroads_open, deer_creek_open, msb_open, needed, summary_table,
    cases_to_pick, inbound_pallets, notes, oc_alert_text=None,
    board_pulls_remaining=None, board_picks_remaining=None,
):
    client = get_groq_client()
    if client is None:
        return (
            "Board analysis could not be completed because GROQ_API_KEY is missing. "
            "Add GROQ_API_KEY in Streamlit Cloud Secrets."
        )

    try:
        board_payload   = json.loads(board_text)
        py_summary      = board_payload.get("python_verified_outbound_summary", {})
        py_inbound      = board_payload.get("python_verified_inbound_summary", {})
        actionable_rows = board_payload.get("actionable_outbound_rows", [])
        completed_rows  = board_payload.get("completed_outbound_rows", [])
    except Exception:
        py_summary      = {}
        py_inbound      = {}
        actionable_rows = []
        completed_rows  = []

    staffing_lines = []
    for task, row in summary_table.iterrows():
        staffing_lines.append(
            f"  {task}: need {int(row['Needed'])}, have {int(row['Assigned'])}, "
            f"gap {int(row['Difference'])} ({row['Status']})"
        )
    staffing_summary = "\n".join(staffing_lines)

    plants_open = [
        p for p, s in [("Crossroads", crossroads_open), ("Deer Creek", deer_creek_open), ("MSB", msb_open)]
        if s == "YES"
    ]

    oc_section = f"\n{oc_alert_text}\n" if oc_alert_text else ""

    # ── Board actual remaining work (from K2/L2) ────────────────────────────
    if board_pulls_remaining is not None and board_picks_remaining is not None:
        actual_remaining_block = (
            f"\nACTUAL REMAINING WORK ON BOARD RIGHT NOW (source of truth — read directly from board totals):\n"
            f"  Pulls remaining: {board_pulls_remaining:,}\n"
            f"  Picks remaining: {board_picks_remaining:,}\n"
            f"  These are live totals. When a load goes RTL or Completed the crew erases its pulls/picks,\n"
            f"  so this reflects only loads that still have work to do.\n"
            f"  Use these numbers — not the projected cases_to_pick — as the basis for all remaining\n"
            f"  workload analysis, short risk assessment, and labor move recommendations.\n"
        )
        # Comparison to projection
        pulls_per_pick = 70  # approximate cases per pull
        board_cases_equivalent = (board_picks_remaining or 0) + (board_pulls_remaining or 0) * pulls_per_pick
        actual_remaining_block += (
            f"  Approx case equivalent (picks + pulls×70): {board_cases_equivalent:,}\n"
            f"  Projected cases_to_pick from inputs: {cases_to_pick:,.0f}\n"
        )
        if cases_to_pick > 0:
            pct = (board_cases_equivalent / cases_to_pick) * 100
            if pct > 110:
                actual_remaining_block += f"  → BEHIND PLAN: board shows {pct:.0f}% of projected workload still remaining.\n"
            elif pct < 90:
                actual_remaining_block += f"  → AHEAD OF PLAN: board shows only {pct:.0f}% of projected workload remaining.\n"
            else:
                actual_remaining_block += f"  → ON TRACK: board shows {pct:.0f}% of projected workload remaining.\n"
    else:
        actual_remaining_block = (
            "\nACTUAL REMAINING WORK: Board totals not available (CSV upload or empty board). "
            "Use projected cases_to_pick as the workload estimate.\n"
        )

    # ── Second-shift goal block ──────────────────────────────────────────────
    if shift == "2nd":
        shift_goal_block = (
            "\n2ND SHIFT MISSION:\n"
            "2nd shift owns everything 1st shift didn't finish. The goal is to get every load on today's board\n"
            "out the door — no exceptions. Key priorities in order:\n"
            "  1. Finish all picks and pulls so no load ships short or misses its appointment.\n"
            "  2. Load and release every live load first — trailers sitting at the dock cost money.\n"
            "  3. Work through drop loads in appointment-time order.\n"
            "  4. Set up the DC so 1st shift tomorrow starts clean: all putaway done, replenishment staged,\n"
            "     dock clear, and any known shorts communicated to the incoming manager.\n"
            "  5. Do not leave picks or pulls sitting in the system — if it's on the board it gets done.\n"
            "When analyzing staffing gaps, labor moves, and priorities — always frame the answer around\n"
            "what 2nd shift needs to do to close out the day completely.\n"
        )
    else:
        shift_goal_block = (
            "\n1ST SHIFT MISSION:\n"
            "1st shift sets the tone. The goal is to get as far ahead as possible so 2nd shift\n"
            "inherits a clean, manageable board. Key priorities:\n"
            "  1. Prevent shorts — replenishment and picking must stay ahead of appointments.\n"
            "  2. Protect live loads — these are at the dock and must leave on time.\n"
            "  3. Clear the early appointment loads first; every RTL by its appt time is a win.\n"
            "  4. Use any extra labor proactively — get picks and pulls done for loads due tonight.\n"
            "  5. Hand off to 2nd shift with the board well-organized: shorts identified, RTLs staged,\n"
            "     and no surprises.\n"
        )

    loads_by_day = py_summary.get("loads_by_day", {})
    day_str = ", ".join(f"{d}:{n}" for d, n in loads_by_day.items())
    verified_counts = (
        f"VERIFIED COUNTS (Python — do not recount):\n"
        f"Total:{py_summary.get('loads_read_from_board',0)}  "
        f"Late:{py_summary.get('late_loads',0)}  "
        f"RTL:{py_summary.get('rtl_loads',0)}  "
        f"R/S:{py_summary.get('rs_loads',0)}  "
        f"Picking:{py_summary.get('picking_loads',0)}  "
        f"Pick/Short:{py_summary.get('picking_short_loads',0)}  "
        f"LoadedShort:{py_summary.get('loaded_short_loads',0)}  "
        f"Completed:{py_summary.get('completed_loads',0)}  "
        f"Blank:{py_summary.get('blank_or_not_started_loads',0)}\n"
        f"Live:{py_summary.get('live_loads',0)}  "
        f"CPU:{py_summary.get('cpu_loads',0)}  "
        f"Canadian:{py_summary.get('canadian_loads',0)}  "
        f"TT4:{py_summary.get('tt4_needed_loads',0)}  "
        f"LoadCheck:{py_summary.get('load_check_loads',0)}  "
        f"LoaderAssigned:{py_summary.get('loads_with_loader_assigned',0)}  "
        f"MissingLoader:{py_summary.get('loads_missing_loader',0)}\n"
        f"By day: {day_str}"
    )

    ib_day_str = ", ".join(f"{d}:{n}" for d, n in py_inbound.get("loads_by_day", {}).items())
    verified_inbound = (
        f"VERIFIED INBOUND COUNTS (Python — do not recount):\n"
        f"Total:{py_inbound.get('loads_read_from_inbound',0)}  "
        f"Live:{py_inbound.get('live_loads',0)}  "
        f"Drop:{py_inbound.get('drop_loads',0)}  "
        f"OnLot:{py_inbound.get('on_lot',0)}  "
        f"AtDoor:{py_inbound.get('at_door',0)}  "
        f"OnLotOrAtDoor:{py_inbound.get('on_lot',0)+py_inbound.get('at_door',0)}\n"
        f"By day: {ib_day_str}"
    )

    actionable_table = _rows_to_table(
        actionable_rows,
        ["day","load","customer","time","door","status","type","pulls","picks","loader","flags","comments"]
    )
    completed_table = _rows_to_table(
        completed_rows,
        ["day","load","customer","time","status"]
    )

    prompt = f"""You are an outbound warehouse shift manager. Data comes from Excel cells — treat it as accurate. Use short bullets. No corporate fluff.

CONTEXT: High-volume grocery DC. 1st shift 06:00-16:30. 2nd shift 16:30-03:00. 24-hr clock. First shift loads ~52% of day's loads.
Priorities: 1)Prevent shorts 2)Protect departures 3)Picking flow 4)Inbound flow 5)Proactive labor use.
Statuses: RTL=fully picked and pulled, staged ready to load (zero remaining work)|R/S=short on full pallets|Picking/Short=inventory shortage|LoadedShort=trailer loaded but missing product, severe service risk|Late=missed/at risk|Live=trailer at dock, highest priority over drop trailers.
Flags: LOAD-CHECK=yellow|TT4-NEEDED=blue|CANADIAN=red font.
Rates: Pick=185 cases/hr/person|Load=1 trailer/hr/person|Unload=44 pallets/hr|Tasking=25 pallets/hr|Ticket avg=60 cases.
Labor rules: Keep pickers picking. Tasking protects pickers. Protect loading labor. Lead/Extra used proactively.
PULLS vs PICKS: Pulls = full pallets pulled from reserve. Picks = individual case picks. Both are remaining work units on that load.
  A load with pulls=0 and picks=0 is done (RTL or erased by crew). Non-zero = work still needed.
{shift_goal_block}
TODAY: {day} {shift} shift | {total_cases:,} total cases scheduled | Projected {cases_to_pick:,.0f} cases to pick this shift (from inputs) | {hours_remaining}hrs left | {total_outbound_loads} loads today | Plants open: {", ".join(plants_open) if plants_open else "none"} | Notes: {notes.strip() or "none"}
{actual_remaining_block}
STAFFING (Python-computed vs inputs projection):
{staffing_summary}

{verified_counts}
{verified_inbound}
{oc_section}
ACTIONABLE LOADS — includes actual pulls and picks remaining per load:
{actionable_table}

COMPLETED LOADS (pacing reference — appt times show when they were due):
{completed_table}

===== OUTPUT — 6 sections =====

1. BOARD SUMMARY
- Loads by status using verified counts. Do not recount.
- Pacing: completed vs total, ahead/on track/behind based on appt times and hours left.
- Late loads: load#, door, what's happening.
- Actual remaining work: reference the board totals (pulls/picks remaining) as the real workload picture.
- Inbound: use VERIFIED INBOUND COUNTS only.

2. OC ALERTS
- Every OC load: load#, customer, status, appt time, exact required actions.
- If none: "No OC customers on today's board."

3. PICKING & SHORT RISK
- Base all remaining work analysis on actual board pulls/picks remaining — NOT the projected cases_to_pick.
- Use projected cases_to_pick only to show whether we are ahead or behind plan.
- Which loads still have pulls/picks and need protection. Which are the biggest risks.
- Short risk assessment: specific loads, specific numbers, specific appt times at risk.
- Labor moves: from where to where, with specific people counts.

4. PRIORITIZATION
- Which load#s to prioritize and why. Reference actual pulls/picks remaining per load.
- Order by urgency: live loads first, then priority group 1, then by appt time.

5. STAFFING CROSS-ANALYSIS
- Compare actual board workload to staffing. Are we resourced correctly for what is actually left?
- What can we fix now? Where does labor move first?
- Achievable shift goal with specific numbers and times.

6. TOP ACTION ITEMS
- Next 30 min: 3 items.
- Next 2 hrs: 3 items.

RULES: Every labor move = from X to Y. Use specific times. Only use board data — never invent loads or numbers. OC alerts must be complete. For 2nd shift: frame every recommendation around finishing the day completely.
"""

    try:
        response = client.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[{"role": "user", "content": prompt}],
            temperature=0.2,
            max_completion_tokens=2000,
        )
        return response.choices[0].message.content

    except Exception as e:
        return f"Board analysis could not be completed: {str(e)}"


def write_board_analysis_to_excel(wb, analysis_text, oc_matches=None):
    sheet_name = "Board Analysis"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(sheet_name)

    dark_blue    = "0F5B78"
    orange       = "C55A11"
    white        = "FFFFFF"
    light_blue   = "D9EAF7"
    light_orange = "FCE4D6"
    thin   = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = "Board Excel Analysis — AI Insights"
    ws["A1"].font      = Font(size=16, bold=True, color=white)
    ws["A1"].fill      = PatternFill("solid", fgColor=dark_blue)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A1:G1")
    ws.row_dimensions[1].height = 28

    ws["A2"] = "Generated by Groq AI — cross-referenced with today's staffing and demand data"
    ws["A2"].font = Font(italic=True, size=10)
    ws["A2"].fill = PatternFill("solid", fgColor=light_blue)
    ws.merge_cells("A2:G2")

    current_row = 4

    if oc_matches:
        ws.cell(current_row, 1).value = "OPPORTUNITY CUSTOMER ALERT — SPECIAL HANDLING REQUIRED"
        ws.cell(current_row, 1).font      = Font(size=13, bold=True, color=white)
        ws.cell(current_row, 1).fill      = PatternFill("solid", fgColor=orange)
        ws.cell(current_row, 1).alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        for match in oc_matches:
            c = match["customer"]
            oc_lines = [
                f"CUSTOMER: {c['name'].upper()}  |  Priority: {c['priority']}",
                f"Issue History: {c['issue']}",
                f"DC Requirements: {c['requirements']}",
            ]
            if c["sign_off"]:
                oc_lines.append("DC Supervisor Sign-Off REQUIRED before this load ships.")
            if c["pictures"]:
                oc_lines.append("Photos REQUIRED: 3 on dock + 3 during loading (6 total). Email to manager.")
            for line in oc_lines:
                cell = ws.cell(current_row, 1, line)
                cell.font      = Font(size=10, bold=("CUSTOMER:" in line))
                cell.fill      = PatternFill("solid", fgColor=light_orange)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border    = border
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
                ws.row_dimensions[current_row].height = max(15, min(60, len(line) // 5))
                current_row += 1
            current_row += 1
        current_row += 1

    for line in analysis_text.split("\n"):
        cell = ws.cell(current_row, 1, line)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border    = border
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
        ws.row_dimensions[current_row].height = max(15, min(60, len(line) // 5))
        current_row += 1

    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 22
    ws.column_dimensions["A"].width = 110


def write_recommendations_to_excel(wb, staff, shift):
    if shift == "1st":
        sheet_name = "Staffing sheet 1ST Shift"
    else:
        sheet_name = "Staffing Sheet 2nd Shift"

    ws_staff = wb[sheet_name]
    for excel_row, task in zip(range(2, len(staff) + 2), staff["Recommended Task"]):
        ws_staff[f"I{excel_row}"] = task

    shift_number = 1 if shift == "1st" else 2
    ws_crew = wb["Crew Sheet"]

    crew_name_to_row = {}
    for row in range(2, ws_crew.max_row + 1):
        name       = ws_crew.cell(row, 1).value
        crew_shift = ws_crew.cell(row, 2).value
        if name and str(crew_shift).strip() == str(shift_number):
            crew_name_to_row[str(name).strip().lower()] = row

    for _, row in staff.iterrows():
        name = str(row["Name"]).strip().lower()
        task = row["Recommended Task"]
        if name in crew_name_to_row:
            crew_row = crew_name_to_row[name]
            ws_crew[f"C{crew_row}"] = task
            ws_crew[f"D{crew_row}"] = task


def build_dashboard(wb, summary_table, present_recommendations, recommendations, oc_matches=None):
    if "Staffing Dashboard" in wb.sheetnames:
        ws_dash = wb["Staffing Dashboard"]
        ws_dash.delete_rows(1, ws_dash.max_row)
    else:
        ws_dash = wb.create_sheet("Staffing Dashboard")

    dark_blue  = "0F5B78"
    orange     = "C55A11"
    light_blue = "D9EAF7"
    green      = "C6EFCE"
    red        = "FFC7CE"
    yellow     = "FFEB9C"
    white      = "FFFFFF"
    thin   = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws_dash["A1"] = "1st Shift Staffing Dashboard"
    ws_dash["A1"].font      = Font(size=18, bold=True, color=white)
    ws_dash["A1"].fill      = PatternFill("solid", fgColor=dark_blue)
    ws_dash["A1"].alignment = Alignment(horizontal="center")
    ws_dash.merge_cells("A1:K1")

    total_present  = len(present_recommendations)
    total_needed   = int(summary_table["Needed"].sum())
    total_assigned = int(summary_table["Assigned"].sum())
    lead_extra     = int((present_recommendations["Recommended Task"] == "Lead/Extra").sum())
    overall_gap    = total_assigned - total_needed

    kpis     = [
        ("Total Present",  total_present),
        ("Total Needed",   total_needed),
        ("Total Assigned", total_assigned),
        ("Lead/Extra",     lead_extra),
        ("Overall Gap",    overall_gap),
    ]
    kpi_cols = [1, 3, 5, 7, 9]
    for (label, value), col in zip(kpis, kpi_cols):
        ws_dash.cell(3, col).value     = label
        ws_dash.cell(4, col).value     = value
        ws_dash.cell(3, col).font      = Font(bold=True, color=white)
        ws_dash.cell(3, col).fill      = PatternFill("solid", fgColor=dark_blue)
        ws_dash.cell(3, col).alignment = Alignment(horizontal="center")
        ws_dash.cell(4, col).font      = Font(bold=True, size=14)
        ws_dash.cell(4, col).fill      = PatternFill("solid", fgColor=light_blue)
        ws_dash.cell(4, col).alignment = Alignment(horizontal="center")
        ws_dash.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col + 1)
        ws_dash.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 1)

    oc_banner_row = 6
    if oc_matches:
        customer_names = ", ".join(m["customer"]["name"].upper() for m in oc_matches)
        ws_dash.cell(oc_banner_row, 1).value = (
            f"OC ALERT: Opportunity Customers on today's board — {customer_names} — See 'Board Analysis' tab for full requirements."
        )
        ws_dash.cell(oc_banner_row, 1).font      = Font(bold=True, color=white, size=11)
        ws_dash.cell(oc_banner_row, 1).fill      = PatternFill("solid", fgColor=orange)
        ws_dash.cell(oc_banner_row, 1).alignment = Alignment(horizontal="center", wrap_text=True)
        ws_dash.merge_cells(start_row=oc_banner_row, start_column=1, end_row=oc_banner_row, end_column=11)
        ws_dash.row_dimensions[oc_banner_row].height = 22
        summary_label_row = oc_banner_row + 2
    else:
        summary_label_row = oc_banner_row

    ws_dash.cell(summary_label_row, 1).value = "Needed vs Assigned"
    ws_dash.cell(summary_label_row, 1).font  = Font(size=14, bold=True)

    header_row = summary_label_row + 1
    headers = ["Task", "Needed", "Assigned", "Difference", "Status"]
    for c, header in enumerate(headers, 1):
        cell           = ws_dash.cell(header_row, c)
        cell.value     = header
        cell.font      = Font(bold=True, color=white)
        cell.fill      = PatternFill("solid", fgColor=dark_blue)
        cell.border    = border
        cell.alignment = Alignment(horizontal="center")

    for r, (task, row) in enumerate(summary_table.iterrows(), header_row + 1):
        values = [task, int(row["Needed"]), int(row["Assigned"]), int(row["Difference"]), row["Status"]]
        for c, value in enumerate(values, 1):
            cell           = ws_dash.cell(r, c)
            cell.value     = value
            cell.border    = border
            cell.alignment = Alignment(horizontal="center")
            if c == 5:
                if value == "Good":
                    cell.fill = PatternFill("solid", fgColor=green)
                elif value == "Understaffed":
                    cell.fill = PatternFill("solid", fgColor=red)
                else:
                    cell.fill = PatternFill("solid", fgColor=yellow)

    ws_dash.cell(summary_label_row, 7).value = "Written Recommendations / What-Ifs"
    ws_dash.cell(summary_label_row, 7).font  = Font(size=14, bold=True)

    rec_row = header_row
    for rec in recommendations:
        ws_dash.cell(rec_row, 7).value     = f"• {rec}"
        ws_dash.cell(rec_row, 7).alignment = Alignment(wrap_text=True, vertical="top")
        ws_dash.merge_cells(start_row=rec_row, start_column=7, end_row=rec_row, end_column=11)
        rec_row += 1

    board_start = max(header_row + len(summary_table) + 4, rec_row + 2)
    ws_dash.cell(board_start, 1).value = "Recommended Staffing Board"
    ws_dash.cell(board_start, 1).font  = Font(size=14, bold=True)

    board_headers = ["Name", "Skills", "Best Fit", "Recommended Task"]
    for c, header in enumerate(board_headers, 1):
        cell           = ws_dash.cell(board_start + 1, c)
        cell.value     = header
        cell.font      = Font(bold=True, color=white)
        cell.fill      = PatternFill("solid", fgColor=dark_blue)
        cell.border    = border
        cell.alignment = Alignment(horizontal="center")

    for r, (_, row) in enumerate(present_recommendations.iterrows(), board_start + 2):
        values = [row["Name"], row["Skills"], row["Best Fit"], row["Recommended Task"]]
        for c, value in enumerate(values, 1):
            cell           = ws_dash.cell(r, c)
            cell.value     = value
            cell.border    = border
            cell.alignment = Alignment(horizontal="center")
            if r % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=light_blue)

    chart_anchor_row = board_start + len(present_recommendations) + 5

    bar = BarChart()
    bar.title        = "Needed vs Assigned"
    bar.y_axis.title = "Workers"
    bar.x_axis.title = "Task"
    data = Reference(ws_dash, min_col=2, max_col=3, min_row=header_row, max_row=header_row + len(summary_table))
    cats = Reference(ws_dash, min_col=1, min_row=header_row + 1, max_row=header_row + len(summary_table))
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    bar.height = 9
    bar.width  = 15
    bar.legend.position = "r"
    ws_dash.add_chart(bar, f"E{chart_anchor_row}")

    pie = PieChart()
    pie.title = "Assigned Labor Distribution"
    pie_data = Reference(ws_dash, min_col=3, min_row=header_row, max_row=header_row + len(summary_table))
    pie_cats = Reference(ws_dash, min_col=1, min_row=header_row + 1, max_row=header_row + len(summary_table))
    pie.add_data(pie_data, titles_from_data=True)
    pie.set_categories(pie_cats)
    pie.height = 9
    pie.width  = 13
    pie.legend.position = "r"
    ws_dash.add_chart(pie, f"I{chart_anchor_row}")

    for col in range(1, 12):
        ws_dash.column_dimensions[get_column_letter(col)].width = 18
    ws_dash.column_dimensions["A"].width = 22
    for col in ["G", "H", "I", "J", "K"]:
        ws_dash.column_dimensions[col].width = 35
    ws_dash.freeze_panes = f"A{header_row}"


def build_email_draft(
    day, shift, total_cases, hours_remaining, total_outbound_loads_day,
    summary_table, present_recommendations, recommendations,
    board_analysis_text=None, oc_matches=None,
    board_pulls_remaining=None, board_picks_remaining=None,
):
    total_present  = len(present_recommendations)
    total_needed   = int(summary_table["Needed"].sum())
    total_assigned = int(summary_table["Assigned"].sum())
    overall_gap    = total_assigned - total_needed
    subject        = f"{day} {shift} Shift Staffing Report"

    staffing_lines = []
    for task, row in summary_table.iterrows():
        staffing_lines.append(
            f"- {task}: Need {int(row['Needed'])}, Assigned {int(row['Assigned'])}, "
            f"Gap {int(row['Difference'])} ({row['Status']})"
        )
    top_recommendations = "\n".join([f"- {rec}" for rec in recommendations[:8]])

    oc_email_block = ""
    if oc_matches:
        oc_names = ", ".join(
            f"{m['customer']['name'].upper()} [{m['customer']['priority']}]"
            for m in oc_matches
        )
        oc_email_block = (
            f"\nOPPORTUNITY CUSTOMER ALERT:\n"
            f"The following OC customers have loads on today's board: {oc_names}.\n"
            f"See the attached staffing report (Board Analysis tab) for full handling requirements.\n"
        )

    board_work_block = ""
    if board_pulls_remaining is not None and board_picks_remaining is not None:
        board_work_block = (
            f"\nActual Remaining Work (from board):\n"
            f"- Pulls remaining: {board_pulls_remaining:,}\n"
            f"- Picks remaining: {board_picks_remaining:,}\n"
        )

    body = f"""
Good morning,

Here is the staffing report for {day} {shift} shift.

Daily Inputs:
- Total cases: {total_cases:,}
- Total outbound loads: {total_outbound_loads_day}
- Hours remaining: {hours_remaining}
- Total present: {total_present}
- Total needed: {total_needed}
- Total assigned: {total_assigned}
- Overall labor gap: {overall_gap}
{board_work_block}
Staffing Summary:
{chr(10).join(staffing_lines)}
{oc_email_block}
Key Recommendations / What-Ifs:
{top_recommendations}
"""

    if board_analysis_text:
        body += f"""

Board Analysis:
{board_analysis_text}
"""
    body += """

The full staffing report is attached.

Thanks,
"""
    return subject, body.strip()


# ============================================================
#  STREAMLIT INTERFACE
# ============================================================

st.sidebar.header("Daily Inputs")

day = st.sidebar.selectbox(
    "Day",
    ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"],
)

shift = st.sidebar.selectbox("Shift", ["1st", "2nd"])

names = load_names_for_shift(shift)

total_cases = st.sidebar.number_input("Total Cases for Today", min_value=0, step=1, value=0)

hours_remaining = st.sidebar.number_input("Hours Remaining in Shift", min_value=0.0, step=0.25, value=8.0)

total_outbound_loads_day = st.sidebar.number_input("Total Outbound Loads for the Day", min_value=0, step=1, value=0)

crossroads_open = st.sidebar.selectbox("Crossroads plant open?", ["YES", "NO"])
deer_creek_open = st.sidebar.selectbox("Deer Creek plant open?", ["YES", "NO"])
msb_open        = st.sidebar.selectbox("MSB plant open?", ["YES", "NO"])

present_workers = st.sidebar.multiselect("Who is present?", names)

notes = st.sidebar.text_area("Operations Notes")

st.markdown("---")
st.subheader("Outbound Board Excel / CSV")

board_file = st.file_uploader(
    "Upload the outbound load board Excel or CSV file",
    type=["xlsx", "xls", "csv"],
    help="Cell values and color flags (yellow = load check, light-blue = TT4, red font = Canadian) are read directly from the file.",
)

if board_file:
    st.success("Board file loaded — ready for analysis.")

    with st.expander("Preview: What Python parsed from the board (no AI tokens used)", expanded=False):
        try:
            board_file.seek(0)
            file_name_lower = board_file.name.lower()
            if file_name_lower.endswith(".csv"):
                preview_rows = board_records_from_csv(board_file)
            else:
                preview_rows = board_records_from_excel(board_file)

            if not preview_rows:
                st.warning("No load rows were parsed.")
            else:
                st.metric("Staff Present Today", len(present_workers))
                st.markdown("---")

                preview_summary = build_python_board_summary(preview_rows)

                col1, col2, col3, col4, col5 = st.columns(5)
                col1.metric("Total Loads",   preview_summary["loads_read_from_board"])
                col2.metric("RTL",           preview_summary["rtl_loads"])
                col3.metric("Picking/Short", preview_summary["picking_short_loads"])
                col4.metric("R/S",           preview_summary["rs_loads"])
                col5.metric("Loaded Short",  preview_summary["loaded_short_loads"])

                col6, col7, col8, col9, col10 = st.columns(5)
                col6.metric("Picking",           preview_summary["picking_loads"])
                col7.metric("Blank/Not Started", preview_summary["blank_or_not_started_loads"])
                col8.metric("Live Loads",        preview_summary["live_loads"])
                col9.metric("CPU Loads",         preview_summary["cpu_loads"])
                col10.metric("Late",             preview_summary["late_loads"])

                st.caption(f"Outbound loads by day: {preview_summary['loads_by_day']}")

                # ── Show board totals from K2/L2 ────────────────────────────
                board_file.seek(0)
                preview_pulls, preview_picks = read_board_totals(board_file)
                if preview_pulls is not None or preview_picks is not None:
                    st.markdown("---")
                    st.markdown("**Actual Remaining Work (board totals K2/L2)**")
                    bt1, bt2 = st.columns(2)
                    bt1.metric("Pulls Remaining", f"{preview_pulls:,}" if preview_pulls else "—")
                    bt2.metric("Picks Remaining", f"{preview_picks:,}" if preview_picks else "—")
                    st.caption(
                        "These are live totals from the board. When a load goes RTL or Completed "
                        "the crew erases its K/L values, so this reflects only loads still needing work."
                    )

                board_file.seek(0)
                inbound_preview_rows = board_records_from_inbound_sheet(board_file)
                if inbound_preview_rows:
                    inbound_preview_summary = build_python_inbound_summary(inbound_preview_rows)
                    st.markdown("---")
                    st.markdown("**Inbound**")
                    ib1, ib2, ib3, ib4 = st.columns(4)
                    ib1.metric("Total Inbound",    inbound_preview_summary["loads_read_from_inbound"])
                    ib2.metric("Live",             inbound_preview_summary["live_loads"])
                    ib3.metric("Drop",             inbound_preview_summary["drop_loads"])
                    ib4.metric("On Lot / At Door", inbound_preview_summary["on_lot"] + inbound_preview_summary["at_door"])
                    st.caption(f"Inbound loads by day: {inbound_preview_summary['loads_by_day']}")
                    inbound_df = pd.DataFrame([
                        {
                            "Day":      r.get("day", ""),
                            "Load #":   r.get("load_number", ""),
                            "Carrier":  r.get("carrier", ""),
                            "Time":     r.get("appt_time", ""),
                            "Type":     r.get("type", ""),
                            "Trailer":  r.get("trailer", ""),
                            "Status":   r.get("status", ""),
                            "Receiver": r.get("receiver", ""),
                            "Origin":   r.get("origin", ""),
                            "Notes":    r.get("notes", ""),
                        }
                        for r in inbound_preview_rows
                    ])
                    st.dataframe(inbound_df, use_container_width=True, height=250)

                st.markdown("---")
                st.markdown("**Every outbound load row Python extracted from the file:**")
                preview_df = pd.DataFrame([
                    {
                        "Day":      r.get("day", ""),
                        "Date":     r.get("date", ""),
                        "Load #":   r.get("load_number", ""),
                        "Customer": r.get("customer", ""),
                        "Carrier":  r.get("carrier", ""),
                        "Time":     r.get("appt_time", ""),
                        "Door":     r.get("door", ""),
                        "Trailer":  r.get("trailer", ""),
                        "Status":   r.get("status", "") or "—",
                        "Type":     r.get("type", ""),
                        "TT4":      r.get("tt4", ""),
                        "Loader":   r.get("loader", ""),
                        "Pulls":    r.get("pulls", 0),
                        "Picks":    r.get("picks", 0),
                        "Flags":    ", ".join(r.get("flags", [])),
                        "Comments": r.get("comments", ""),
                    }
                    for r in preview_rows
                ])
                st.dataframe(preview_df, use_container_width=True, height=400)

                st.markdown("**Quick sanity checks:**")
                issues = []
                blank_time = [r["load_number"] for r in preview_rows if not r.get("appt_time")]
                if blank_time:
                    issues.append(f" {len(blank_time)} load(s) have no time parsed: {', '.join(blank_time[:5])}{'...' if len(blank_time) > 5 else ''}")
                blank_customer = [r["load_number"] for r in preview_rows if not r.get("customer")]
                if blank_customer:
                    issues.append(f" {len(blank_customer)} load(s) have no customer name: {', '.join(blank_customer[:5])}")
                no_day = [r["load_number"] for r in preview_rows if not r.get("day")]
                if no_day:
                    issues.append(f" {len(no_day)} load(s) have no day context: {', '.join(no_day[:5])}")
                if issues:
                    for issue in issues:
                        st.warning(issue)
                else:
                    st.success("All loads have time, customer, and day context — parse looks clean.")

        except Exception as e:
            st.error(f"Preview failed: {e}")
            st.exception(e)

with st.expander("View Opportunity Customer List (from Excel file)"):
    oc_list_preview = load_oc_customer_list()
    if oc_list_preview:
        oc_preview_rows = []
        for c in oc_list_preview:
            oc_preview_rows.append({
                "Customer":          c["name"].title(),
                "Customer #":        c["customer_number"] or "—",
                "Priority":          c["priority"],
                "Issue":             c["issue"],
                "DC Requirements":   c["requirements"],
                "Sign-Off Required": "Yes" if c["sign_off"] else "No",
                "Photos Required":   "Yes" if c["pictures"] else "No",
            })
        st.dataframe(pd.DataFrame(oc_preview_rows), use_container_width=True)
        st.caption(f"Loaded {len(oc_list_preview)} customers from '{OC_FILE}'")
    else:
        st.warning(f"No customers loaded. Check that '{OC_FILE}' exists in the app folder.")

st.markdown("---")

if st.button("Generate Staffing Report"):
    working_file = f"working_staffing_file_{day}_{shift}.xlsx"
    shutil.copyfile(TEMPLATE_FILE, working_file)

    wb = load_workbook(working_file)
    ws = wb["Inputs"]

    total_outbound_loads_actual = total_outbound_loads_day * 0.52

    ws["B1"] = day
    ws["B2"] = shift
    ws["B3"] = total_cases
    ws["B4"] = hours_remaining
    ws["B8"] = crossroads_open
    ws["B9"] = deer_creek_open
    ws["B10"] = msb_open

    cases_to_pick, full_pallets = calculate_input_values(day, shift, total_cases)
    ws["B5"] = cases_to_pick
    ws["B6"] = full_pallets
    ws["B7"] = total_outbound_loads_actual

    selected = {name.strip().lower() for name in present_workers}
    ws_crew_ref = wb["Crew Sheet"]

    crew_name_to_inputs_row = {}
    for _r in range(2, ws_crew_ref.max_row + 1):
        crew_name = ws_crew_ref.cell(_r, 1).value
        if crew_name and str(crew_name).strip():
            crew_name_to_inputs_row[str(crew_name).strip().lower()] = _r + 1

    for _r in range(3, max(crew_name_to_inputs_row.values(), default=3) + 1):
        ws.cell(_r, 7).value = ""

    for worker in selected:
        if worker in crew_name_to_inputs_row:
            ws.cell(crew_name_to_inputs_row[worker], 7).value = "x"

    ws["B12"] = notes

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc  = True
    wb.save(working_file)

    needed, raw_needed, cases_to_pick, full_pallets, inbound_pallets = calculate_needed(
        day, shift, total_cases, hours_remaining, total_outbound_loads_actual,
        crossroads_open, deer_creek_open, msb_open,
    )

    if shift == "1st":
        staffing_sheet = "Staffing sheet 1ST Shift"
    else:
        staffing_sheet = "Staffing Sheet 2nd Shift"

    staff = pd.read_excel(working_file, sheet_name=staffing_sheet, usecols="A,D,F,H,I")
    staff.columns = ["Name", "Skills", "Best Fit", "Present", "Recommended Task"]
    staff = staff[staff["Name"].notna()].copy()

    staff["Present"] = staff["Name"].astype(str).str.strip().str.lower().apply(
        lambda x: "x" if x in selected else ""
    )

    staff = generate_recommendations(staff, needed)
    present_recommendations, summary_table = build_summary(staff, needed)
    recommendations = build_recommendations(
        summary_table, present_recommendations, raw_needed, hours_remaining, notes
    )

    wb = load_workbook(working_file)
    write_recommendations_to_excel(wb, staff, shift)

    board_analysis_text   = None
    oc_matches            = []
    board_pulls_remaining = None
    board_picks_remaining = None

    if board_file is not None:
        # Read K2/L2 board totals first — independent of the AI parse
        board_file.seek(0)
        board_pulls_remaining, board_picks_remaining = read_board_totals(board_file)

        with st.spinner("Reading board file → scanning for Opportunity Customers → running AI analysis..."):
            board_text    = read_board_file_to_text(board_file)
            oc_matches    = find_oc_customers_in_board(board_text)
            oc_alert_text = build_oc_alert_text(oc_matches)

            if oc_matches:
                customer_names_found = [m["customer"]["name"].upper() for m in oc_matches]
                st.warning(
                    f"**Opportunity Customer Alert:** "
                    f"The following customers were detected on today's board and require special handling: "
                    f"**{', '.join(customer_names_found)}**. "
                    f"See the OC Alerts section below for full requirements."
                )

            board_analysis_text = analyze_board_with_groq(
                board_text=board_text,
                day=day,
                shift=shift,
                total_cases=total_cases,
                hours_remaining=hours_remaining,
                total_outbound_loads=total_outbound_loads_day,
                crossroads_open=crossroads_open,
                deer_creek_open=deer_creek_open,
                msb_open=msb_open,
                needed=needed,
                summary_table=summary_table,
                cases_to_pick=cases_to_pick,
                inbound_pallets=inbound_pallets,
                notes=notes,
                oc_alert_text=oc_alert_text,
                board_pulls_remaining=board_pulls_remaining,
                board_picks_remaining=board_picks_remaining,
            )

            write_board_analysis_to_excel(wb, board_analysis_text, oc_matches=oc_matches)

    build_dashboard(wb, summary_table, present_recommendations, recommendations, oc_matches=oc_matches)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    try:
        os.remove(working_file)
    except Exception:
        pass

    st.success("Staffing report generated successfully.")

    # ── Show actual remaining work prominently ───────────────────────────────
    if board_pulls_remaining is not None or board_picks_remaining is not None:
        st.markdown("---")
        st.subheader("Actual Remaining Work on Board")
        rw1, rw2 = st.columns(2)
        rw1.metric("Pulls Remaining", f"{board_pulls_remaining:,}" if board_pulls_remaining else "0")
        rw2.metric("Picks Remaining", f"{board_picks_remaining:,}" if board_picks_remaining else "0")

    if oc_matches:
        st.markdown("---")
        st.subheader("Opportunity Customer Alerts")
        st.error(
            "The following customers on today's board are on the **Opportunity Customer List** "
            "and require special DC actions before their loads ship."
        )
        for match in oc_matches:
            c = match["customer"]
            with st.expander(f"{c['name'].upper()}  —  Priority: {c['priority']}", expanded=True):
                st.markdown(f"**Issue History:** {c['issue']}")
                st.markdown(f"**DC Requirements:** {c['requirements']}")
                if c["sign_off"]:
                    st.markdown("**DC Supervisor Sign-Off REQUIRED before this load ships.**")
                if c["pictures"]:
                    st.markdown("**Photos REQUIRED:** 3 on dock + 3 during loading (6 total). Email to manager.")
    elif board_file is not None:
        st.info("No Opportunity Customers detected on today's board.")

    st.subheader("Staffing Summary")
    st.dataframe(summary_table, use_container_width=True)

    st.subheader("Recommended Staffing Board")
    st.dataframe(
        present_recommendations[["Name", "Skills", "Best Fit", "Recommended Task"]].reset_index(drop=True),
        use_container_width=True,
    )

    st.subheader("Written Recommendations / What-Ifs")
    for rec in recommendations:
        st.write(f"• {rec}")

    if board_analysis_text:
        st.markdown("---")
        st.subheader("Board Excel Analysis — AI Insights")
        st.info(
            "The analysis below was generated by Groq AI reading the board Excel/CSV file directly "
            "from cell values, including color flags for load checks, TT4s, and Canadian loads."
        )
        st.markdown(board_analysis_text)

    st.download_button(
        label="Download Staffing Report",
        data=output,
        file_name="Staffing Report Generated.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    email_subject, email_body = build_email_draft(
        day=day,
        shift=shift,
        total_cases=total_cases,
        hours_remaining=hours_remaining,
        total_outbound_loads_day=total_outbound_loads_day,
        summary_table=summary_table,
        present_recommendations=present_recommendations,
        recommendations=recommendations,
        board_analysis_text=board_analysis_text,
        oc_matches=oc_matches,
        board_pulls_remaining=board_pulls_remaining,
        board_picks_remaining=board_picks_remaining,
    )

    st.markdown("---")
    st.subheader("Email Ready to Send")
    st.text_input("Email Subject", value=email_subject)
    st.text_area("Email Body", value=email_body, height=500)
